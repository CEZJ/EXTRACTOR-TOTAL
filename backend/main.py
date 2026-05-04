import os
import re
import shutil
import pdfplumber
import pandas as pd
import numpy as np
from fastapi.responses import FileResponse
from typing import List
from datetime import datetime
from openpyxl import Workbook 
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware

print("Cargando librerias de Inteligencia Artificial... (EasyOCR)")
import easyocr
lector_ocr = easyocr.Reader(['es'], gpu=False, verbose=False)

MESES_TEXTO = {
    'enero': '01', 'febrero': '02', 'marzo': '03', 'abril': '04',
    'mayo': '05', 'junio': '06', 'julio': '07', 'agosto': '08',
    'septiembre': '09', 'setiembre': '09', 'octubre': '10', 'noviembre': '11', 'diciembre': '12'
}

# ==========================================
# FUNCIONES AUXILIARES (TUS FUNCIONES INTACTAS)
# ==========================================
def limpiar_fecha(fecha_str):
    if not fecha_str: return ""
    return re.sub(r'\s+', '', fecha_str)

def buscar_numero_largo(palabras_clave, texto, min_len=4):
    fragmentos = re.finditer(rf"(?:{palabras_clave}).{{0,150}}", texto, re.IGNORECASE | re.DOTALL)
    for fragmento in fragmentos:
        bloque = fragmento.group(0)
        bloque = re.sub(r'\bN[oº]?\b|\bNro\.?\b', '', bloque, flags=re.IGNORECASE)
        candidatos = re.findall(r'\b[A-Z0-9\-]*[0-9]{' + str(min_len) + r',}[A-Z0-9\-]*\b', bloque)
        if candidatos: return candidatos[0]
    return ""

def atrapar_fechas_vigencia(texto, palabras_clave=r'Vigencia de P[oó]liza|Vigencia'):
    fragmentos = re.finditer(rf"(?:{palabras_clave}).{{0,400}}", texto, re.IGNORECASE | re.DOTALL)
    for fragmento in fragmentos:
        fechas = re.findall(r'\b\d{2}\s*[/-]\s*\d{2}\s*[/-]\s*\d{4}\b', fragmento.group(0))
        if len(fechas) >= 2: return limpiar_fecha(fechas[0]), limpiar_fecha(fechas[1])
    return "", ""

def atrapar_fecha_emision(texto, palabras_clave=r'Fecha\s*(?:de\s*)?Emisi[oó]n|Emisi[oó]n|FECHA(?!\s*Vencimiento)'):
    fragmentos = re.finditer(rf'(?:{palabras_clave}).{{0,500}}', texto, re.IGNORECASE | re.DOTALL)
    for fragmento in fragmentos:
        fechas = re.findall(r'\b\d{2}\s*[/-]\s*\d{2}\s*[/-]\s*\d{4}\b', fragmento.group(0))
        if fechas: return limpiar_fecha(fechas[0])
    match_larga = re.search(r'Emitida\s+en.*?(\d{1,2})\s+de\s+([a-zA-Z]+)\s+(?:de|del)\s+(\d{4})', texto, re.IGNORECASE | re.DOTALL)
    if match_larga:
        dia, mes, anio = match_larga.groups()
        return f"{int(dia):02d}/{MESES_TEXTO.get(mes.lower(), '01')}/{anio}"
    return ""

def atrapar_monto_cercano(palabras_clave, texto):
    fragmentos = re.finditer(rf"(?:{palabras_clave}).{{0,500}}", texto, re.IGNORECASE | re.DOTALL)
    for fragmento in fragmentos:
        bloque = fragmento.group(0)
        monto = re.search(r'(?:S/\.|S/|S\s*/|5/|\$|USD|US\$|S\s*-|SOLES)\s*([\d]+(?:[.,]\d{3})*(?:[.,]\d{1,2})?)\b', bloque, re.IGNORECASE)
        if monto: return monto.group(1)
        monto_decimal = re.search(r'\b([\d]+(?:[.,]\d{3})*[.,]\d{1,2})\b', bloque)
        if monto_decimal: return monto_decimal.group(1)
    return ""

def formatear_dolares(monto, texto_completo, palabra_clave):
    if not monto: return ""
    fragmento = re.search(rf"({palabra_clave}.{{0,50}})", texto_completo, re.DOTALL | re.IGNORECASE)
    if fragmento and ("$" in fragmento.group(0) or "USD" in fragmento.group(0).upper()):
        return f"($) {monto}"
    return monto

# ==========================================
# EXTRACCION PRINCIPAL
# ==========================================
def extraer_datos_pdf(ruta_archivo):
    texto_completo = ""
    with pdfplumber.open(ruta_archivo) as pdf:
        paginas_a_procesar = pdf.pages[:10] 
        for pagina in paginas_a_procesar:
            texto = pagina.extract_text()
            if not texto or len(texto.strip()) < 20:
                try:
                    img = pagina.to_image(resolution=200).original
                    img_np = np.array(img)
                    texto = "\n".join(lector_ocr.readtext(img_np, detail=0))
                except Exception:
                    texto = ""
            if texto: texto_completo += texto + "\n"

    datos = {
        "Archivo": os.path.basename(ruta_archivo), "Ruc_DNI": "", "Poliza_Contrato": "",
        "Documento": "", "Vigencia_Inicio": "", "Vigencia_Fin": "", "Fecha_Emision": "",
        "Prima_Total": "", "Fecha_pago": ""
    }
    texto_upper = texto_completo.upper()

    if "PROTECTA" in texto_upper:
        datos["Ruc_DNI"] = buscar_numero_largo(r"DNI/RUC:|RUC", texto_completo, 8)
        datos["Poliza_Contrato"] = buscar_numero_largo(r"P[oó]liza|Contrato", texto_completo, 4)
        doc = ""
        m_sctr = re.search(r"(?:AC|PF|CS)\s*-\s*SCTR\s*-\s*(\d+)", texto_completo, re.IGNORECASE)
        if m_sctr: doc = m_sctr.group(1)
        if not doc:
            m_fact = re.search(r"F\d{3}\s*-\s*(\d+)", texto_completo)
            if m_fact and re.search(r"FACTURA|SOAT", texto_completo, re.IGNORECASE): doc = m_fact.group(1)
        if not doc:
            if re.search(r"C[oó]digo\s*SBS[\s:]*VI", texto_completo, re.IGNORECASE) or "PENSIONES" in texto_upper: doc = "Buscar en plataforma"
        if not doc:
            doc_sucio = re.search(r"AVISO DE COBRANZA.*?FECHA.{0,30}?([A-Z0-9\-]{5,})", texto_completo, re.DOTALL)
            doc = re.sub(r"[A-Za-z\-]+", "", doc_sucio.group(1)) if doc_sucio else ""
        if not doc: doc = "Buscar en plataforma"
        datos["Documento"] = doc
        datos["Vigencia_Inicio"], datos["Vigencia_Fin"] = atrapar_fechas_vigencia(texto_completo, r"Vigencia")
        if not datos["Vigencia_Inicio"] or not datos["Vigencia_Fin"]:
            m_vig = re.search(r'Vigencia\s*[:.-]?\s*(?:Del\s*)?(\d{2}\s*[/-]\s*\d{2}\s*[/-]\s*\d{4})\s*al\s*(\d{2}\s*[/-]\s*\d{2}\s*[/-]\s*\d{4})', texto_completo, re.IGNORECASE)
            if m_vig:
                datos["Vigencia_Inicio"] = limpiar_fecha(m_vig.group(1))
                datos["Vigencia_Fin"] = limpiar_fecha(m_vig.group(2))
        emision = ""
        fe_match = re.search(r'Fecha\s*de\s*Emisi[oó]n', texto_completo, re.IGNORECASE)
        if fe_match:
            ventana = texto_completo[max(0, fe_match.start() - 80):fe_match.end() + 100]
            fecha_encontrada = re.search(r'\d{2}\s*[/\-]\s*\d{2}\s*[/\-]\s*\d{4}', ventana)
            if fecha_encontrada: emision = limpiar_fecha(fecha_encontrada.group(0))
        if not emision: emision = atrapar_fecha_emision(texto_completo, r"Fecha\s*(?:de\s*)?Emisi[oó]n|Emisi[oó]n|FECHA(?!\s*Vencimiento)")
        if not emision and datos["Documento"] == "Buscar en plataforma": emision = datos.get("Vigencia_Inicio", "")
        datos["Fecha_Emision"] = emision
        monto = ""
        m_importe = re.search(r'Importe\s*Total[\s:.-]+(?:S/\.|S/|5/|\$|S\s*-)?\s*([\d]+(?:[.,]\d{3})*(?:[.,]\d{1,2})?)', texto_completo, re.IGNORECASE)
        if m_importe: monto = m_importe.group(1)
        if not monto:
            for kw in [r"Prima Comercial Total m[aá]s IGV", r"PRIMA COMERCIAL TOTAL", r"Importe Total", r"TOTAL A PAGAR", r"PRIMA TOTAL"]:
                monto = atrapar_monto_cercano(kw, texto_completo)
                if monto: break
        if not monto:
            m_monto = re.search(r'Importe\s*Total\s*[:.-]?\s*(?:S/|S/\.|USD|\$|S\s*/)?\s*([\d]+(?:[.,]\d{3})*(?:[.,]\d{1,2})?)', texto_completo, re.IGNORECASE)
            if m_monto: monto = m_monto.group(1)
        datos["Prima_Total"] = formatear_dolares(monto, texto_completo, "PRIMA|Importe|TOTAL")

    elif "PACIFICO" in texto_upper or "PACÍFICO" in texto_upper:
        datos["Ruc_DNI"] = buscar_numero_largo(r"R\.?U\.?C\.?", texto_completo, 11) or "No aplica"
        m_tabla_gen = re.search(r'\b(\d{6,9})\s+(\d{5,9})\s+\d{2,10}\s+\d{2}\s*[/-]\s*\d{2}\s*[/-]\s*\d{4}', texto_completo)
        poliza = ""
        m_pol_exacto = re.search(r'P[oó]liza/Contrato:\s*([A-Z0-9\-]{5,})', texto_completo, re.IGNORECASE)
        if m_pol_exacto: poliza = m_pol_exacto.group(1)
        if not poliza:
            m_pol_tabla = re.search(r'(?:AC|F\d{3})[\s\-]*\d+\s+(\d{6,})', texto_completo)
            if m_pol_tabla: poliza = m_pol_tabla.group(1)
        if not poliza and m_tabla_gen: poliza = m_tabla_gen.group(2)
        if not poliza: poliza = buscar_numero_largo(r"P[OÓ]LIZA(?:/Contrato)?|Contrato", texto_completo, 5)
        datos["Poliza_Contrato"] = poliza
        doc = ""
        m_doc_acob = re.search(r'A/COB\s*Giro.{0,100}?\b(\d{7,10})\b', texto_completo, re.IGNORECASE | re.DOTALL)
        if m_doc_acob: doc = m_doc_acob.group(1)
        if not doc:
            m_doc_tabla = re.search(r'(?:AC|F\d{3})[\s\-]*(\d{5,})', texto_completo)
            if m_doc_tabla: doc = m_doc_tabla.group(1)
        if not doc and m_tabla_gen: doc = m_tabla_gen.group(1)
        if not doc: doc = buscar_numero_largo(r"A/COB|LIQUIDACION DE PRIMA|Aviso de Cobranza", texto_completo, 5)
        datos["Documento"] = doc
        datos["Vigencia_Inicio"], datos["Vigencia_Fin"] = atrapar_fechas_vigencia(texto_completo, r"Vigencia")
        datos["Fecha_Emision"] = atrapar_fecha_emision(texto_completo, r"Fecha\s*(?:de\s*)?Emisi[oó]n|Emisi[oó]n")
        prima = ""
        m_monto_tabla = re.search(r'\d{2}\s*[/-]\s*\d{2}\s*[/-]\s*\d{4}[\s\-]+\d{2}\s*[/-]\s*\d{2}\s*[/-]\s*\d{4}\s+([\d]+(?:[.,]\d{3})*(?:[.,]\d{2}))', texto_completo)
        if m_monto_tabla: prima = m_monto_tabla.group(1)
        if not prima:
            palabras_monto_pacifico = r"Prima\s*Comercial\s*\+\s*INT.*?IGV|Importe\s*Total|TOTAL\s*A\s*COBRAR|Total\s*a\s*Pagar|PRIMA\s*TOTAL"
            monto = atrapar_monto_cercano(palabras_monto_pacifico, texto_completo)
            prima = formatear_dolares(monto, texto_completo, "IGV|TOTAL|Importe|Pagar")
            if not prima:
                monto_alt = re.search(rf'(?:{palabras_monto_pacifico}).{{0,800}}?(?:S/\.|S/|5/|\$|USD|US\$|S\s*-)?\s*([\d]+(?:[.,]\d{{3}})*(?:[.,]\d{{2}})?)', texto_completo, re.IGNORECASE | re.DOTALL)
                if monto_alt: prima = formatear_dolares(monto_alt.group(1), texto_completo, "IGV|TOTAL|COBRAR")
        datos["Prima_Total"] = prima
        if not datos["Fecha_Emision"]:
            emision_alt = re.search(r'(?:Fecha\s*(?:de\s*)?Emisi[oó]n|Emisi[oó]n).{0,800}?(\d{2}\s*[/-]\s*\d{2}\s*[/-]\s*\d{4})', texto_completo, re.IGNORECASE | re.DOTALL)
            if emision_alt: datos["Fecha_Emision"] = limpiar_fecha(emision_alt.group(1))

    elif "SANITAS" in texto_upper or "CRECER" in texto_upper:
        datos["Ruc_DNI"] = buscar_numero_largo(r"DNI/RUC|RUC", texto_completo, 8)
        datos["Poliza_Contrato"] = buscar_numero_largo(r"Contrato", texto_completo, 4)
        doc = ""
        m_sctr_san = re.search(r'(?:PF|CS)\s*-\s*SCTR\s*-\s*(\d+)', texto_completo, re.IGNORECASE)
        if m_sctr_san: doc = m_sctr_san.group(1)
        if not doc:
            m_fact_san = re.search(r"F\d{3}\s*-\s*(\d+)", texto_completo)
            if m_fact_san and "FACTURA" in texto_upper: doc = m_fact_san.group(1)
        if not doc:
            doc_sucio = re.search(r"(?:PROFORMA|AVISO DE COBRANZA).*?FECHA.{0,30}?([A-Z0-9\-]{5,})", texto_completo, re.DOTALL | re.IGNORECASE)
            doc = re.sub(r"[A-Za-z\-]+", "", doc_sucio.group(1)) if doc_sucio else ""
        datos["Documento"] = doc
        datos["Vigencia_Inicio"], datos["Vigencia_Fin"] = atrapar_fechas_vigencia(texto_completo, r"Vigencia")
        datos["Fecha_Emision"] = atrapar_fecha_emision(texto_completo, r"Fecha de Emisi[oó]n|FECHA(?!\s*Vencimiento)")
        monto = ""
        m_monto_san = re.search(r'(?:Importe\s*Total|PRECIO\s*VENTA\s*TOTAL|PRIMA\s*TOTAL)[\s\S]{1,300}?(?:S/|S/\.|USD|\$|S\s*/)?\s*(?<![/-])(?<!\d)([\d]+(?:[.,]\d{3})*(?:[.,]\d{1,2})?)(?![/-])(?!\d)', texto_completo, re.IGNORECASE)
        if m_monto_san: monto = m_monto_san.group(1)
        if not monto: monto = atrapar_monto_cercano(r"Importe Total|TOTAL", texto_completo)
        datos["Prima_Total"] = formatear_dolares(monto, texto_completo, "TOTAL|Importe Total")

    elif "MAPFRE" in texto_upper:
        datos["Ruc_DNI"] = buscar_numero_largo(r"RUC", texto_completo, 11)
        poliza = ""
        m_pol_exacto = re.search(r'P[OÓ0-9A-Z]*LIZA[^\d]{0,40}?\b(\d{8,})\b', texto_completo, re.IGNORECASE)
        if m_pol_exacto: poliza = m_pol_exacto.group(1)
        if not poliza:
            m_pol_sbs = re.search(r'\b[A-Z]{2}\d{8,10}\b[\s\S]{1,200}?\b(\d{13,15})\b', texto_completo)
            if m_pol_sbs: poliza = m_pol_sbs.group(1)
        if not poliza:
            pol_candidata = buscar_numero_largo(r"P[OÓ]LIZA", texto_completo, 4)
            if pol_candidata and pol_candidata.isdigit(): poliza = pol_candidata
        datos["Poliza_Contrato"] = poliza
        doc = ""
        m_doc_recibo = re.search(r'(?:NRO[.\s]*RECIBO|CRONOGRAMA DE PAGO)[\s\S]{1,800}?(?<!\d)(\d{8,11})(?!\d)', texto_completo, re.IGNORECASE)
        if m_doc_recibo: doc = m_doc_recibo.group(1)
        if not doc:
            doc_sucio = re.search(r"DOC\. IDENTIFIC\..{0,20}?(?:[A-Z]+\s+)?([A-Z0-9]{5,})", texto_completo)
            doc_candidato = doc_sucio.group(1) if doc_sucio else buscar_numero_largo(r"RECIBO", texto_completo, 5)
            if doc_candidato and not "FECHA" in doc_candidato.upper(): doc = doc_candidato
        datos["Documento"] = doc
        datos["Vigencia_Inicio"], datos["Vigencia_Fin"] = atrapar_fechas_vigencia(texto_completo, r"VIGENCIA")
        emision = ""
        m_emi = re.search(r'EMISI[OÓ0-9A-Z]*N[^\d]{0,40}?(\d{2}\s*[/-]\s*\d{2}\s*[/-]\s*\d{4})', texto_completo, re.IGNORECASE)
        if m_emi: emision = limpiar_fecha(m_emi.group(1))
        if not emision:
            m_emi_sbs = re.search(r'\b[A-Z]{2}\d{8,10}\b[\s\S]{1,80}?\b(\d{2}\s*[/-]\s*\d{2}\s*[/-]\s*\d{4})\b', texto_completo)
            if m_emi_sbs: emision = limpiar_fecha(m_emi_sbs.group(1))
        datos["Fecha_Emision"] = emision
        prima = ""
        m_prima_igv = re.search(r'Prima\s*Comercial\s*\+(?:\s*I[\.\s]*G[\.\s]*V)?[\.\s]*([\d]+(?:[.,]\d{3})*(?:[.,]\d{1,2})?)', texto_completo, re.IGNORECASE)
        if m_prima_igv: prima = m_prima_igv.group(1)
        if not prima and doc:
            m_cron_monto = re.search(rf'{doc}[\s\S]{{1,300}}?([\d]+(?:[.,]\d{{3}})*(?:[.,]\d{{1,2}})?)', texto_completo, re.IGNORECASE)
            if m_cron_monto: prima = m_cron_monto.group(1)
        if not prima:
            monto = atrapar_monto_cercano(r"Prima\s*Comercial\s*\+(?:\s*I\.?G\.?V)?|TOTAL|Importe Total", texto_completo)
            prima = formatear_dolares(monto, texto_completo, "TOTAL|Prima Comercial|Importe Total")
        datos["Prima_Total"] = prima

    elif "RIMAC" in texto_upper or "RÍMAC" in texto_upper:
        datos["Ruc_DNI"] = buscar_numero_largo(r"R\.?U\.?C\.?", texto_completo, 11)
        poliza = buscar_numero_largo(r"POLIZA SEG|P[óo]liza", texto_completo, 4)
        if poliza and "-" in poliza: poliza = poliza.split("-")[-1].strip()
        datos["Poliza_Contrato"] = poliza
        doc = ""
        m_liq = re.search(r'Liquidaci[oó]n[\s\w]*?(?:Nro\.?|Prima)[\s:N°º]*(\d{7,})', texto_completo, re.IGNORECASE)
        if m_liq: doc = m_liq.group(1)
        if not doc: doc = buscar_numero_largo(r"Documento de Cobranza|Documento.*?LQ|LQ", texto_completo, 5)
        datos["Documento"] = doc
        datos["Vigencia_Inicio"], datos["Vigencia_Fin"] = atrapar_fechas_vigencia(texto_completo, r"Vigencia de P[oó]liza|Vigencia")
        emision = atrapar_fecha_emision(texto_completo, r"Emisi[oó]n")
        if not emision:
            match_fecha = re.search(r'(\d{1,2})\s+de\s+([a-zA-Z]+)\s+(?:de|del)\s+(\d{4})', texto_completo, re.IGNORECASE)
            if match_fecha:
                dia, mes, anio = match_fecha.groups()
                emision = f"{int(dia):02d}/{MESES_TEXTO.get(mes.lower(), '01')}/{anio}"
        datos["Fecha_Emision"] = emision if emision else datos.get("Vigencia_Inicio", "")
        prima = ""
        m_prima_rimac = re.search(r'(?:Prima\s*Comercial\s*\+\s*IGV|Incluye\s*TCEA\s*e\s*IGV)[\s:S/.]*([\d]+(?:[.,]\d{3})*(?:[.,]\d{1,2})?)', texto_completo, re.IGNORECASE)
        if m_prima_rimac: prima = m_prima_rimac.group(1)
        if not prima:
            monto = atrapar_monto_cercano(r"TOTAL\s*A\s*PAGAR|Prima Comercial Total|IMPORTE TOTAL", texto_completo)
            prima = formatear_dolares(monto, texto_completo, "Prima Comercial|TOTAL|IMPORTE")
        datos["Prima_Total"] = prima

    elif "LA POSITIVA" in texto_upper or "POSITIVA" in texto_upper:
        datos["Ruc_DNI"] = "20605619453"
        datos["Poliza_Contrato"] = buscar_numero_largo(r"P[óo]liza", texto_completo, 4)
        datos["Documento"] = buscar_numero_largo(r"Proforma", texto_completo, 5)
        vig_ini = atrapar_fecha_emision(texto_completo, r"Vigencia[\s\-]*Inicio|Desde")
        vig_fin = atrapar_fecha_emision(texto_completo, r"T[eé]rmino|Hasta")
        if vig_ini and vig_fin:
            datos["Vigencia_Inicio"], datos["Vigencia_Fin"] = vig_ini, vig_fin
        else:
            datos["Vigencia_Inicio"], datos["Vigencia_Fin"] = atrapar_fechas_vigencia(texto_completo, r"Vigencia")
        match_fecha = re.search(r'(\d{1,2})\s+de\s+([a-zA-Z]+)\s+(?:de|del)\s+(\d{4})', texto_completo, re.IGNORECASE)
        if match_fecha:
            dia, mes, anio = match_fecha.groups()
            datos["Fecha_Emision"] = f"{int(dia):02d}/{MESES_TEXTO.get(mes.lower(), '01')}/{anio}"
        monto = atrapar_monto_cercano(r"Prima Comercial|TOTAL", texto_completo)
        datos["Prima_Total"] = formatear_dolares(monto, texto_completo, "Prima|TOTAL")

    if not datos["Vigencia_Inicio"]:
        ini_alt = re.search(r'Desde\s*[:.-]?\s*(\d{2}\s*[/-]\s*\d{2}\s*[/-]\s*\d{4})', texto_completo, re.IGNORECASE)
        if ini_alt: datos["Vigencia_Inicio"] = limpiar_fecha(ini_alt.group(1))
    if not datos["Vigencia_Fin"]:
        fin_alt = re.search(r'(?:Hasta|al)\s*[:.-]?\s*(\d{2}\s*[/-]\s*\d{2}\s*[/-]\s*\d{4})', texto_completo, re.IGNORECASE)
        if fin_alt: datos["Vigencia_Fin"] = limpiar_fecha(fin_alt.group(1))

    for clave, valor in datos.items():
        if not valor or str(valor).strip() == "":
            if clave == "Documento": datos[clave] = "No aplica"

    if "SANITAS" in texto_upper or "PROTECTA" in texto_upper or "CRECER" in texto_upper:
        datos["Fecha_pago"] = datos["Vigencia_Inicio"]
    else:
        datos["Fecha_pago"] = datos["Fecha_Emision"]

    return datos

# ==========================================
# FORMATO EXCEL Y TRAMA MASIVA
# ==========================================
def aplicar_formato_excel(ruta_excel):
    wb = load_workbook(ruta_excel)
    ws = wb.active
    color_cabecera = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    color_fila_par = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    color_alerta   = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    fuente_blanca  = Font(color="FFFFFF", bold=True)
    fuente_normal  = Font(color="000000", bold=False)

    for celda in ws[1]:
        celda.fill = color_cabecera
        celda.font = fuente_blanca
        celda.alignment = Alignment(horizontal="center", vertical="center")

    for fila in range(2, ws.max_row + 1):
        es_par = (fila % 2 == 0)
        for celda in ws[fila]:
            if celda.value == "Buscar en plataforma":
                celda.fill, celda.font = color_alerta, fuente_blanca
            else:
                celda.fill = color_fila_par if es_par else PatternFill()
                celda.font = fuente_normal

    ws.auto_filter.ref = ws.dimensions
    for col in ws.columns:
        max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2
    wb.save(ruta_excel)

def generar_trama_masiva(resultados, ruta_directorio):
    directorio_script = os.path.dirname(os.path.abspath(__file__))
    ruta_plantilla = os.path.join(directorio_script, "trama_carga_masiva.xlsx")
    
    if not os.path.exists(ruta_plantilla): 
        print(f"\n❌ [ERROR CRÍTICO]: No encuentro tu plantilla base en: {ruta_plantilla}")
        print("Asegúrate de tener el archivo 'trama_carga_masiva.xlsx' en esa carpeta.\n")
        return
    # 1. Creamos el Excel desde cero
    wb = Workbook()

    # --- SELLO INVISIBLE DE AUTORÍA ---
    wb.properties.creator = "Carlos Enrique Zegarra Jurado - PROPIEDAD INTELECTUAL"
    wb.properties.title = "C.Z.A.R. Engine Extraction"
    
    ws = wb.active
    ws.title = "TRAMA"
    
    # 2. LOS ENCABEZADOS EXACTOS DE TU PLANTILLA
    cabeceras = [
        "POLIZA_CERTF", "AVISO_COB_NUM_PRIMERA_CUOTA", "TIPO_DOC", "TIPO_PAGO", 
        "AVISO_VIGENCIA_INICIO", "AVISO_VIGENCIA_FIN", "FECHA_EMISION", 
        "FECHA_SOLICITUD_SEGURO", "FECHA_ENTREGA_POLIZA_CLIENTE", "PRIMA_NETA", 
        "PRIMA_TOTAL", "NUMERO_CUOTAS", "IMPORTE_CADA_CUOTA", 
        "FECHA_PRIMER_VENCIMIENTO", "FECHA_PAGO", "MOTIVO"
    ]
    ws.append(cabeceras)
    
    # Damos formato elegante a tu cabecera
    color_cabecera = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    fuente_blanca = Font(color="FFFFFF", bold=True)
    for celda in ws[1]:
        celda.fill = color_cabecera
        celda.font = fuente_blanca
        celda.alignment = Alignment(horizontal="center", vertical="center")

    def string_a_fecha(fecha_str):
        if not fecha_str: return None
        try: return datetime.strptime(fecha_str.replace('-', '/'), "%d/%m/%Y")
        except ValueError: return None

    # 3. Rellenamos con los datos y fórmulas
    fila = 2 
    for datos in resultados:
        ws[f'A{fila}'] = datos.get("Poliza_Contrato", "")
        
        doc_original = str(datos.get("Documento", ""))
        match_sctr = re.search(r'SCTR\s*-\s*(\d+)', doc_original, re.IGNORECASE)
        ws[f'B{fila}'] = match_sctr.group(1) if match_sctr else doc_original
        
        ws[f'C{fila}'] = f'=IF(K{fila}<0, "DEVOLUCION", IF(K{fila}<60, "ENDOSO", "RENOVACION"))'
        ws[f'D{fila}'] = "CONTADO"
        
        fechas_map = {'E': "Vigencia_Inicio", 'F': "Vigencia_Fin", 'G': "Fecha_Emision", 'O': "Fecha_pago"}
        for col, key in fechas_map.items():
            obj_fecha = string_a_fecha(datos.get(key, ""))
            if obj_fecha:
                ws[f'{col}{fila}'].value = obj_fecha
                ws[f'{col}{fila}'].number_format = 'dd/mm/yyyy'

        for col_f in ['H', 'I', 'N']:
            ws[f'{col_f}{fila}'] = f'=G{fila}' if col_f != 'N' else f'=G{fila}+30'
            ws[f'{col_f}{fila}'].number_format = 'dd/mm/yyyy'
        
        prima_str = str(datos.get("Prima_Total", "0")).replace("($)", "").replace(",", "").strip()
        ws[f'K{fila}'] = float(prima_str) if prima_str else 0.0
        ws[f'J{fila}'] = f'=ROUND(K{fila}/1.18, 2)'
        ws[f'L{fila}'] = 1
        ws[f'M{fila}'] = f'=K{fila}/L{fila}'
        ws[f'P{fila}'] = f'=UPPER(TEXT(E{fila}, "MMMM"))'
        fila += 1

    # Ajustar el ancho de las columnas automáticamente para que no se vea amontonado
    for col in ws.columns:
        max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    # 4. Guardamos la trama final en tu carpeta de salidas
    wb.save(os.path.join(ruta_directorio, "trama_carga_masiva_FINAL.xlsx"))

# ==========================================
# INICIALIZACIÓN DE LA API FASTAPI
# ==========================================
app = FastAPI(title="Extractor de PDFs API")

# Configuración de CORS para conectar con React
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.post("/procesar-pdfs/")
async def procesar_pdfs_lote(archivos: List[UploadFile] = File(...)):
    os.makedirs("temp_pdfs", exist_ok=True)
    directorio_base = os.path.dirname(os.path.abspath(__file__))
    
    resultados = []
    
    # 1. Extraer los datos de TODOS los PDFs enviados
    for file in archivos:
        ruta_archivo = os.path.join("temp_pdfs", file.filename)
        with open(ruta_archivo, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
            
        try:
            datos = extraer_datos_pdf(ruta_archivo)
            resultados.append(datos)
        except Exception as e:
            print(f"Error con {file.filename}: {e}")
        finally:
            if os.path.exists(ruta_archivo):
                os.remove(ruta_archivo)

# 2. Aplicar TU LÓGICA EXACTA (Generar solo 2 archivos en carpeta aparte)
    if resultados:
        # --- NUEVO: Creamos una carpeta para que no se mezcle con tu código ---
        carpeta_resultados = os.path.join(directorio_base, "Archivos_Excel")
        os.makedirs(carpeta_resultados, exist_ok=True)
        
        # Archivo 1: Reporte_Polizas.xlsx (Ahora dentro de la subcarpeta)
        df = pd.DataFrame(resultados)[[
            "Archivo", "Ruc_DNI", "Poliza_Contrato", "Documento",
            "Vigencia_Inicio", "Vigencia_Fin", "Fecha_Emision", "Prima_Total", "Fecha_pago"
        ]]
        
        # CAMBIO AQUÍ: Usamos carpeta_resultados en lugar de directorio_base
        ruta_excel = os.path.join(carpeta_resultados, "Reporte_Polizas.xlsx")
        df.to_excel(ruta_excel, index=False)
        aplicar_formato_excel(ruta_excel)

        # Archivo 2: trama_carga_masiva_FINAL.xlsx (También a la subcarpeta)
        # CAMBIO AQUÍ: Enviamos la nueva ruta a tu función
        generar_trama_masiva(resultados, carpeta_resultados)

        return {"status": "success", "mensaje": "Lote procesado. Archivos guardados en subcarpeta.", "datos": resultados}

# ==========================================
# ENDPOINTS PARA DESCARGA DE ARCHIVOS
# ==========================================

@app.get("/descargar-reporte/")
async def descargar_reporte():
    directorio_base = os.path.dirname(os.path.abspath(__file__))
    # Apuntamos a la subcarpeta "Archivos_Excel"
    ruta = os.path.join(directorio_base, "Archivos_Excel", "Reporte_Polizas.xlsx")
    if os.path.exists(ruta):
        return FileResponse(path=ruta, filename="Reporte_Polizas.xlsx")
    raise HTTPException(status_code=404, detail="No encontrado")

@app.get("/descargar-trama/")
async def descargar_trama():
    directorio_base = os.path.dirname(os.path.abspath(__file__))
    # Apuntamos a la subcarpeta "Archivos_Excel"
    ruta = os.path.join(directorio_base, "Archivos_Excel", "trama_carga_masiva_FINAL.xlsx")
    if os.path.exists(ruta):
        return FileResponse(path=ruta, filename="trama_carga_masiva_FINAL.xlsx")
    raise HTTPException(status_code=404, detail="No encontrado")

if __name__ == "__main__":
    import uvicorn
    import os
    port = int(os.environ.get("PORT", 8000))
    uvicorn.run(app, host="0.0.0.0", port=port)